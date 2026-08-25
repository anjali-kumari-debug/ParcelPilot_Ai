"""Load the assessment workbook into SQLite.

Sheets: README (metadata), accounts, orders, tickets.
We create one table per data sheet plus an `actions` table where the agent's
(mocked) state-changing actions are recorded after user confirmation.

Run standalone:  python -m app.data.ingest_xlsx
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from .. import config

# Sheets that become tables. README is metadata only (not a table).
DATA_SHEETS = ["accounts", "orders", "tickets"]


def _clean(value) -> str | None:
    """Normalise a cell into a string (or None). Keeps SQLite columns simple/TEXT."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v if v != "" else None
    return str(value)


def _ingest_sheet(conn: sqlite3.Connection, ws) -> int:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    headers = [str(h).strip() for h in rows[0] if h is not None]
    ncols = len(headers)

    # Recreate table each run so ingestion is idempotent.
    conn.execute(f'DROP TABLE IF EXISTS "{ws.title}"')
    cols_sql = ", ".join(f'"{h}" TEXT' for h in headers)
    conn.execute(f'CREATE TABLE "{ws.title}" ({cols_sql})')

    placeholders = ", ".join(["?"] * ncols)
    insert_sql = f'INSERT INTO "{ws.title}" VALUES ({placeholders})'

    count = 0
    for raw in rows[1:]:
        # Skip fully-empty rows (openpyxl pads short rows with None).
        cells = [_clean(c) for c in list(raw)[:ncols]]
        cells += [None] * (ncols - len(cells))
        if all(c is None for c in cells):
            continue
        conn.execute(insert_sql, cells)
        count += 1
    return count


def _read_readme(ws) -> dict[str, str]:
    meta: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] is not None:
            key = str(row[0]).strip()
            val = _clean(row[1]) if len(row) > 1 else None
            if key and val is not None:
                meta[key] = val
    return meta


def ingest_xlsx(xlsx_path: Path | None = None, db_path: Path | None = None) -> dict:
    xlsx_path = xlsx_path or (config.DOC_DIR / "ParcelPilot_Assessment_Data.xlsx")
    db_path = db_path or config.DB_PATH
    db_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(str(xlsx_path), data_only=True)
    summary: dict[str, int] = {}

    conn = sqlite3.connect(str(db_path))
    try:
        for sheet in DATA_SHEETS:
            if sheet in wb.sheetnames:
                summary[sheet] = _ingest_sheet(conn, wb[sheet])

        # Metadata table (dataset snapshot etc.) for reference/traceability.
        if "README" in wb.sheetnames:
            meta = _read_readme(wb["README"])
            conn.execute("DROP TABLE IF EXISTS dataset_meta")
            conn.execute("CREATE TABLE dataset_meta (key TEXT, value TEXT)")
            for k, v in meta.items():
                conn.execute("INSERT INTO dataset_meta VALUES (?, ?)", (k, v))

        # Actions table: append-only record of confirmed, mocked actions.
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS actions (
                action_id   TEXT PRIMARY KEY,
                action_type TEXT,
                account_id  TEXT,
                target_id   TEXT,
                payload     TEXT,
                created_by  TEXT,
                created_at  TEXT
            )
            """
        )
        conn.commit()
    finally:
        conn.close()

    return summary


if __name__ == "__main__":
    result = ingest_xlsx()
    print("xlsx -> SQLite complete:", result)
